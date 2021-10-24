# !pip install openpyxl
import re
import openpyxl
import API1
from API1 import user
import time
import pandas as pd
from pandas import DataFrame
from elasticsearch import Elasticsearch, helpers


es = Elasticsearch()
es2 = Elasticsearch("http://103.74.122.196:9200")




def evalAges(user, dictObj):
    ages = ''
    birthYear = user['infor']['birthYear']
    if 1920 <= birthYear <= 1955:
        ages = '>65'
    elif 1956 <= birthYear <= 1965:
        ages = '55-65'
    elif 1966 <= birthYear <= 1975:
        ages = '45-54'
    elif 1976 <= birthYear <= 1985:
        ages = '35-44'
    elif 1986 <= birthYear <= 1990:
        ages = '30-34'
    elif 1991 <= birthYear <= 1997:
        ages = '23-29'
    elif 1998 <= birthYear <= 2002:
        ages = '18-22'
    elif 2003 <= birthYear <= 2020:
        ages = '<18'
    if ages == '':
        try:
            predictionAges = user['infor']['prediction']['ages']['scores']
            maxProbabilityAges = 0
            for age in predictionAges:
                if predictionAges[age] > maxProbabilityAges:
                    maxProbabilityAges = predictionAges[age]
                    ages = age
        except KeyError:
            pass

    for age in dictObj['Tuổi']:
        strCheck = re.findall('\(\S*', age)[0][1:-1]
        if ages == strCheck:
            if birthYear != '-1':
                dictObj['Tuổi'][age] = birthYear
            else:
                dictObj['Tuổi'][age] = '1'
            return 1
    return 0


def evalRelationship(user, dictObj):
    count = 0
    relation = ''
    relationship = user['infor']['relationship']
    if relationship == 'Góa' or relationship == 'Đã ly hôn' or relationship == 'Đã ly thân':
        relation = 'Ly thân/Ly dị/Goá/Đơn thân'
    elif relationship == 'Hẹn hò' or relationship == 'Đã đính hôn' or relationship == 'Chung sống' or relationship == 'Có mối quan hệ phức tạp':
        relation = 'Hẹn hò'
    elif relationship == 'Đã kết hôn':
        relation = 'Đã kết hôn'
    elif relationship == 'Độc thân' or relationship == 'Tìm hiểu':
        relation = 'Độc thân'

    if relation == '':
        try:
            predictionRelations = user['infor']['prediction']['relations']['scores']
            maxProbabilityRel = 0
            for rel in predictionRelations:
                if predictionRelations[rel] > maxProbabilityRel:
                    maxProbabilityRel = predictionRelations[rel]
                    relation = rel
            relation = 'Độc thân' if relation == 'single' else 'Hẹn hò' if relation == 'in_relationship' \
                else 'Ly thân/Ly dị/Goá/Đơn thân' if relation == 'broken' else 'Đã kết hôn'
        except KeyError:
            pass
    for rel in dictObj['Tình trạng hôn nhân']:
        if relation == rel:
            count += 1
            if relation == 'Đã kết hôn':
                dictObj['Tình trạng hôn nhân'][rel] = 'Đã kết hôn'
            else:
                dictObj['Tình trạng hôn nhân'][rel] = '1'
            return 1
    return 0


def evalGender(user, dicObj):
    gender = user['infor']['gender']
    gender = 'Nam' if gender == 'male' else 'Nữ' if gender == 'female' else 'Khác'
    for sex in dicObj['Giới tính']:
        if sex == gender:
            dicObj['Giới tính'][sex] = '1'
            return 1
    return 0


def evalChild(user, dicObj):
    hasChild = 0
    try:
        hasChild = user['infor']['prediction']['childs']['scores']['has_child']
        dicObj['Gia đình']['Đã có con hay chưa'] = '1' if hasChild > 0.6 else ''
        if dicObj['Gia đình']['Đã có con hay chưa'] == '1': return 1
    except KeyError:
        pass
    return 0


def evalJob(user, dicObj, Sheet):
    count_job = 0
    predictionJob = ''
    listNameGroup = []
    listDescriptionGroup = []
    listPosition = []
    try:
        listPosition = [pos['position'].lower() for pos in user['infor']['works'] if pos['position'] is not None]
    except:
        pass
    try:
        listNameGroup = [name['name'].lower() for name in user['infor_group'] if name['name'] is not None]
        listDescriptionGroup = [name['description'].lower() for name in user['infor_group'] if
                                name['description'] is not None]
    except:
        pass
    try:
        predictionJob = user['infor']['prediction']['job'][0] if len(user['infor']['prediction']['job']) > 0 else None
    except:
        pass

    for i in range(29, 61):
        cellNameC = f'C{i}'
        cellNameJ = f'J{i}'
        cellNameM = f'M{i}'
        cellNameN = f'N{i}'
        if predictionJob == Sheet[cellNameC].value:
            dicObj['Nghề nghiệp'][Sheet[cellNameC].value] = '1'

        listKeyPosition = Sheet[cellNameJ].value.split(', ') if Sheet[cellNameJ].value is not None else []
        listKeyGroup = Sheet[cellNameM].value.split(', ') if Sheet[cellNameM].value is not None else []
        listMust = Sheet[cellNameN].value.split(', ') if Sheet[cellNameN].value is not None else []

        for key in listKeyPosition:
            for pos in listPosition:
                if pos.find(key) != -1:
                    dicObj['Nghề nghiệp'][Sheet[cellNameC].value] = '1'
        count = 0
        for key in listKeyGroup:
            if len(listMust) > 0:
                for m in listMust:
                    for group in listNameGroup:
                        if group.find(key) != -1 and group.find(m) != 1:
                            count += 1
                    for group in listDescriptionGroup:
                        if group.find(key) != -1 and group.find(m) != 1:
                            count += 1
            else:
                for group in listNameGroup:
                    if group.find(key) != -1:
                        count += 1
                for group in listDescriptionGroup:
                    if group.find(key) != -1:
                        count += 1
        dicObj['Nghề nghiệp'][Sheet[cellNameC].value] = '1' if count >= 3 else ''
        if dicObj['Nghề nghiệp'][Sheet[cellNameC].value] == '1': count_job += 1
    return count_job


def evalEdu(user, dicObj, Sheet):
    count_edu = 0
    predictEdu = ''
    listEdu = []
    try:
        predictEdu = user['infor']['prediction']['educationDegree'].lower()
    except:
        pass
    try:
        listEdu = [edu['school'].lower() for edu in user['infor']['educations'] if edu['school'] is not None]
    except:
        pass
    listEdu.append(predictEdu)

    highest_edu = ''

    for i in range(61, 66):
        cellNameC = f'C{i}'
        cellNameM = f'M{i}'

        listKeyEdu = Sheet[cellNameM].value.split(', ') if Sheet[cellNameM].value is not None else []
        for key in listKeyEdu:
            for edu in listEdu:
                if edu.find(key) != -1:
                    highest_edu = Sheet[cellNameC].value
        # if dicObj['trình độ học vấn'][Sheet[cellNameC].value] == '1': count_edu += 1
    if highest_edu != '':
        dicObj['trình độ học vấn'][highest_edu] = '1'
        count_edu += 1
    return count_edu


def evalLanguage(user, dicObj, Sheet):
    count_language = 0
    listLanguage = []
    try:
        listLanguage = [lang.lower() for lang in user['infor']['languages'] if lang is not None]
    except:
        pass

    for i in range(66, 92):
        cellNameC = f'C{i}'
        cellNameM = f'M{i}'
        listKeyLang = Sheet[cellNameM].value.split(', ') if Sheet[cellNameM].value is not None else []
        for key in listKeyLang:
            for lang in listLanguage:
                if lang.find(key) != -1:
                    dicObj['Ngôn ngữ'][Sheet[cellNameC].value] = '1'
        if dicObj['Ngôn ngữ'][Sheet[cellNameC].value] == '1': count_language += 1
    return count_language


def evalHometown(user, dicObj, Sheet):
    listHomeTowns = []
    try:
        listHomeTowns = [town.lower() for town in user['infor']['hometowns'] if town is not None]
    except:
        pass
    listHomeTownProvinces = []
    try:
        listHomeTownProvinces = [town.lower() for town in user['infor']['hometownProvinces'] if town is not None]
    except:
        pass
    for i in range(92, 155):
        cellNameC = f'C{i}'
        cellNameM = f'M{i}'
        listKeyHometown = Sheet[cellNameM].value.split(', ') if Sheet[cellNameM].value is not None else []
        for key in listKeyHometown:
            for town in listHomeTowns:
                if town.find(key) != -1:
                    dicObj['Quê quán'][Sheet[cellNameC].value] = '1'
            for town in listHomeTownProvinces:
                if town.find(key) != -1:
                    dicObj['Quê quán'][Sheet[cellNameC].value] = '1'
        if dicObj['Quê quán'][Sheet[cellNameC].value] == '1':
            return 1

    return 0


def evalLocation(user, dicObj, Sheet):
    listLocation = []
    try:
        listLocation = [town.lower() for town in user['infor']['location'] if town is not None]
    except:
        pass
    listLocationProvinces = []
    try:
        listLocationProvinces = [town.lower() for town in user['infor']['locationProvinces'] if town is not None]
    except:
        pass

    for i in range(155, 218):
        cellNameC = f'C{i}'
        cellNameM = f'M{i}'
        listKeyHometown = Sheet[cellNameM].value.split(', ') if Sheet[cellNameM].value is not None else []
        for key in listKeyHometown:
            for town in listLocation:
                if town.find(key) != -1:
                    dicObj['Nơi ở hiện tại'][Sheet[cellNameC].value] = '1'
            for town in listLocationProvinces:
                if town.find(key) != -1:
                    dicObj['Nơi ở hiện tại'][Sheet[cellNameC].value] = '1'
        if dicObj['Nơi ở hiện tại'][Sheet[cellNameC].value] == '1': return 1
    return 0


def evalOwn(user, dicObj, Sheet):
    count_own = 0
    listNameGroup = []
    listDescriptionGroup = []
    try:
        listNameGroup = [name['name'].lower() for name in user['infor_group'] if name['name'] is not None]
        listDescriptionGroup = [name['description'].lower() for name in user['infor_group'] if
                                name['description'] is not None]
    except:
        pass

    for i in range(220, 225):
        cellNameE = f'E{i}'
        cellNameD = f'D{i}'
        cellNameC = f'C{i}'
        cellNameM = f'M{i}'
        cellNameN = f'N{i}'
        count = 0
        listKeyGroup = Sheet[cellNameM].value.split(', ') if Sheet[cellNameM].value is not None else []
        listMust = Sheet[cellNameN].value.split(', ') if Sheet[cellNameN].value is not None else []
        for key in listKeyGroup:
            if len(listMust) > 0:
                for m in listMust:
                    for group in listNameGroup:
                        if group.find(key) != -1 and group.find(m) != 1:
                            count += 1
                    for group in listDescriptionGroup:
                        if group.find(key) != -1 and group.find(m) != 1:
                            count += 1
            else:
                for group in listNameGroup:
                    if group.find(key) != -1:
                        count += 1
                for group in listDescriptionGroup:
                    if group.find(key) != -1:
                        count += 1
        dicObj['Sở hữu']['Bất động sản'][Sheet[cellNameD].value] = '1' if count >= 2 else ''
        if dicObj['Sở hữu']['Bất động sản'][Sheet[cellNameD].value] == '1': count_own += 1

    for i in range(225, 242):
        cellNameE = f'E{i}'
        cellNameD = f'D{i}'
        cellNameC = f'C{i}'
        cellNameM = f'M{i}'
        cellNameN = f'N{i}'
        count = 0
        listKeyGroup = Sheet[cellNameM].value.split(', ') if Sheet[cellNameM].value is not None else []
        listMust = Sheet[cellNameN].value.split(', ') if Sheet[cellNameN].value is not None else []
        for key in listKeyGroup:
            if len(listMust) > 0:
                for m in listMust:
                    for group in listNameGroup:
                        if group.find(key) != -1 and group.find(m) != 1:
                            count += 1
                    for group in listDescriptionGroup:
                        if group.find(key) != -1 and group.find(m) != 1:
                            count += 1
            else:
                for group in listNameGroup:
                    if group.find(key) != -1:
                        count += 1
                for group in listDescriptionGroup:
                    if group.find(key) != -1:
                        count += 1
        dicObj['Sở hữu']['Bất động sản']['BĐS nổi bật'][Sheet[cellNameE].value] = '1' if count >= 2 else ''
        if dicObj['Sở hữu']['Bất động sản']['BĐS nổi bật'][Sheet[cellNameE].value] == '1': count_own += 1
    for i in range(242, 267):
        cellNameE = f'E{i}'
        cellNameD = f'D{i}'
        cellNameC = f'C{i}'
        cellNameM = f'M{i}'
        cellNameN = f'N{i}'

        count = 0
        listKeyGroup = Sheet[cellNameM].value.split(', ') if Sheet[cellNameM].value is not None else []
        listMust = Sheet[cellNameN].value.split(', ') if Sheet[cellNameN].value is not None else []
        for key in listKeyGroup:
            if len(listMust) > 0:
                for m in listMust:
                    for group in listNameGroup:
                        if group.find(key) != -1 and group.find(m) != 1:
                            count += 1
                    for group in listDescriptionGroup:
                        if group.find(key) != -1 and group.find(m) != 1:
                            count += 1
            else:
                for group in listNameGroup:
                    if group.find(key) != -1:
                        count += 1
                for group in listDescriptionGroup:
                    if group.find(key) != -1:
                        count += 1
        dicObj['Sở hữu']['Xe cộ'][Sheet[cellNameD].value] = '1' if count >= 2 else ''
        if dicObj['Sở hữu']['Xe cộ'][Sheet[cellNameD].value] == '1': count_own += 1
    return count_own


def check_behavior(user, index, keywords, listMust):
    if index == "user_cogroup_index":
        count = 0
        for place in user["infor"]["checkin"]:
            name_place = place["place"]["name"].lower()
            for keyword in keywords:
                if name_place.find(keyword) != -1:
                    count = count + 1
                    break
        if count > 1:
            return 1
        else:
            return 0

    # if index == "dsminer_group":
    #     count = 0
    #     for group in user["infor_group"]:
    #         name = group["name"].lower()
    #         description = group["description"].lower()
    #         if len(listMust) > 0:
    #             for m in listMust:
    #                 for keyword in keywords:
    #                     if (name.find(keyword) != -1 or description.find(keyword) != -1) and (
    #                             name.find(m) != -1 or description.find(m) != -1):
    #                         count = count + 1
    #                         break
    #         else:
    #             for keyword in keywords:
    #                 if name.find(keyword) != -1 or description.find(keyword) != -1:
    #                     count = count + 1
    #                     break
    #     if count > 3:
    #         return 1
    #     else:
    #         return 0
    if index == "post_index":
        count = 0
        for post in user["infor_post"]:
            message = post["message"].lower()
            if len(listMust) > 0:
                for m in listMust:
                    for keyword in keywords:
                        if message.find(keyword) != -1 and message.find(m) != -1:
                            count = count + 1
                            break
            else:
                for keyword in keywords:
                    if message.find(keyword) != -1:
                        count = count + 1
                        break
        if count > 0:
            return 1
        else:
            return 0
    return 0


is_age = 0
is_relationship = 0
is_gender = 0
is_child = 0
is_job = 0
is_edu = 0
is_language = 0
is_hometown = 0
is_location = 0
is_own = 0
is_nganhang = 0
is_baohiem = 0
is_vay = 0
is_the = 0
is_dautu = 0
is_suckhoe = 0
is_giaoduc = 0
is_dulichnuocngoai = 0
is_dulichtrongnuoc = 0
is_duhoc = 0
is_thethao = 0
is_anuong = 0
is_khoahoc = 0
is_nghethuat = 0
is_game = 0
is_sothichsang = 0
is_xeco = 0
is_thoitrang = 0
is_nhac = 0
is_thucpham = 0
is_phim = 0
is_lamdep = 0


class API2:
    def __init__(self):
        pass

    def get_all(self, id, user):
        count_field = 0
        dictObj = {}
        level1 = ''
        level2 = ''
        level3 = ''
        level4 = ''
        level5 = ''
        level6 = ''
        wb = openpyxl.load_workbook('demographic_behaviorKeywords.xlsx')
        Sheet1 = wb['Sheet1']

        check_age = False
        check_gender = False
        check_relation = False
        check_child = False
        check_job = False
        check_edu = False
        check_language = False
        check_hometown = False
        check_location = False
        check_own = False
        for i in range(3, 267):
            cellNameB = f'B{i}'
            cellNameC = f'C{i}'
            cellNameD = f'D{i}'
            cellNameE = f'E{i}'
            cellNameF = f'F{i}'
            cellNameG = f'G{i}'

            cellDataB = Sheet1[cellNameB].value
            cellDataC = Sheet1[cellNameC].value
            cellDataD = Sheet1[cellNameD].value
            cellDataE = Sheet1[cellNameE].value
            cellDataF = Sheet1[cellNameF].value
            cellDataG = Sheet1[cellNameG].value

            if cellDataB is not None:
                level1 = cellDataB
                if cellDataC is None:
                    dictObj[level1] = ''
                else:
                    dictObj[level1] = {}

            if cellDataC is not None:
                level2 = cellDataC
                if cellDataD is None:
                    dictObj[level1][level2] = ''
                else:
                    dictObj[level1][level2] = {}

            if cellDataD is not None:
                level3 = cellDataD
                if cellDataE is None:
                    dictObj[level1][level2][level3] = ''
                else:
                    dictObj[level1][level2][level3] = {}

            if cellDataE is not None:
                level4 = cellDataE
                if cellDataF is None:
                    dictObj[level1][level2][level3][level4] = ''
                else:
                    dictObj[level1][level2][level3][level4] = {}

            if cellDataF is not None:
                level5 = cellDataF
                if cellDataG is None:
                    dictObj[level1][level2][level3][level4][level5] = ''
                else:
                    dictObj[level1][level2][level3][level4][level5] = {}

            if cellDataG is not None:
                level6 = cellDataG
                dictObj[level1][level2][level3][level4][level5][level6] = ''
        if evalAges(user, dictObj):
            global is_age
            is_age += 1
        if evalRelationship(user, dictObj):
            global is_relationship
            is_relationship += 1
        if evalGender(user, dictObj):
            global is_gender
            is_gender += 1
        if evalChild(user, dictObj):
            global is_child
            is_child += 1
        if evalJob(user, dictObj, Sheet1):
            global is_job
            is_job += 1
        if evalEdu(user, dictObj, Sheet1):
            global is_edu
            is_edu += 1
        if evalLanguage(user, dictObj, Sheet1):
            global is_language
            is_language += 1
        if evalHometown(user, dictObj, Sheet1):
            global is_hometown
            is_hometown += 1
        if evalLocation(user, dictObj, Sheet1):
            global is_location
            is_location += 1
        if evalOwn(user, dictObj, Sheet1):
            global is_own
            is_own += 1
        check_nganhang = False
        check_baohiem = False
        check_vay = False
        check_the = False
        check_dautu = False
        check_suckhoe = False
        check_giaoduc = False
        check_dulichtrongnuoc = False
        check_dulichnuocngoai = False
        check_duhoc = False
        check_thethao = False
        check_anuong = False
        check_khoahoc = False
        check_nghethuat = False
        check_game = False
        check_sothichsang = False
        check_xeco = False
        check_thoitrang = False
        check_nhac = False
        check_thucpham = False
        check_phim = False
        check_lamdep = False

        for i in range(268, 1215):
            cellNameB = f'B{i}'
            cellNameC = f'C{i}'
            cellNameD = f'D{i}'
            cellNameE = f'E{i}'
            cellNameF = f'F{i}'
            cellNameG = f'G{i}'
            cellNameIndex1 = f'H{i}'
            cellNameIndex2 = f'K{i}'
            cellNameField1 = f'I{i}'
            cellNameField2 = f'L{i}'
            cellNameKeywords1 = f'J{i}'
            cellNameKeywords2 = f'M{i}'
            cellNameMust = f'N{i}'

            cellDataB = Sheet1[cellNameB].value
            cellDataC = Sheet1[cellNameC].value
            cellDataD = Sheet1[cellNameD].value
            cellDataE = Sheet1[cellNameE].value
            cellDataF = Sheet1[cellNameF].value
            cellDataG = Sheet1[cellNameG].value
            index_1 = Sheet1[cellNameIndex1].value if Sheet1[cellNameIndex1].value is not None else None
            index_2 = Sheet1[cellNameIndex2].value if Sheet1[cellNameIndex2].value is not None else None
            field_1 = Sheet1[cellNameField1].value.split(', ') if Sheet1[cellNameField1].value is not None else None
            field_2 = Sheet1[cellNameField2].value.split(', ') if Sheet1[cellNameField2].value is not None else None
            keywords_1 = Sheet1[cellNameKeywords1].value.split(', ') if Sheet1[
                                                                            cellNameKeywords1].value is not None else None
            keywords_2 = Sheet1[cellNameKeywords2].value.split(', ') if Sheet1[
                                                                            cellNameKeywords2].value is not None else None
            must = Sheet1[cellNameMust].value.split(', ') if Sheet1[cellNameMust].value is not None else []

            value = 0
            if index_1 != None and keywords_1 != None and index_2 == None and keywords_2 == None:
                value = check_behavior(user, index_1, keywords_1, must)
            if index_1 == None and keywords_1 == None and index_2 != None and keywords_2 != None:
                value = check_behavior(user, index_2, keywords_2, must)
            if index_1 != None and keywords_1 != None and index_2 != None and keywords_2 != None:
                value = check_behavior(user, index_2, keywords_2, must) or check_behavior(user, index_1, keywords_1,
                                                                                          must)
            if value == 0:
                value = ""
            if value == 1: value = '1'

            if cellDataB is not None:
                level1 = cellDataB
                if cellDataC is None:
                    dictObj[level1] = value
                else:
                    dictObj[level1] = {}

            if cellDataC is not None:
                level2 = cellDataC
                if cellDataD is None:
                    dictObj[level1][level2] = value
                else:
                    dictObj[level1][level2] = {}

            if cellDataD is not None:
                level3 = cellDataD
                if cellDataE is None:
                    dictObj[level1][level2][level3] = value
                else:
                    dictObj[level1][level2][level3] = {}

            if cellDataE is not None:
                level4 = cellDataE
                if cellDataF is None:
                    dictObj[level1][level2][level3][level4] = value
                else:
                    dictObj[level1][level2][level3][level4] = {}

            if cellDataF is not None:
                level5 = cellDataF
                if cellDataG is None:
                    dictObj[level1][level2][level3][level4][level5] = value
                else:
                    dictObj[level1][level2][level3][level4][level5] = {}

            if cellDataG is not None:
                level6 = cellDataG
                dictObj[level1][level2][level3][level4][level5][level6] = value
            if level1 == 'Ngân hàng' and value == '1':
                check_nganhang = True
            if level1 == 'Bảo hiểm' and value == '1':
                check_baohiem = True
            if level1 == 'vay' and value == '1':
                check_vay = True
            if level1 == 'thẻ' and value == '1':
                check_the = True
            if level1 == 'đầu tư' and value == '1':
                check_dautu = True
            if level1 == 'sức khỏe' and value == '1':
                check_suckhoe = True
            if level1 == 'Giáo dục' and value == '1':
                check_giaoduc = True
            if level1 == 'Du lịch' and level2 == 'Nước ngoài' and value == '1':
                check_dulichnuocngoai = True
            if level1 == 'Du lịch' and level2 == 'Trong nước' and value == '1':
                check_dulichtrongnuoc = True
            if level1 == 'du học' and value == '1':
                check_duhoc = True
            if level2 == 'thể thao' and value == '1':
                check_thethao = True
            if level2 == 'ăn uống' and value == '1':
                check_anuong = True
            if level2 == 'khoa học công nghệ' and value == '1':
                check_khoahoc = True
            if level2 == 'nghệ thuật' and value == '1':
                check_nghethuat = True
            if level2 == 'game' and value == '1':
                check_game = True
            if level2 == 'Sở thích sang' and value == '1':
                check_sothichsang = True
            if level2 == 'xe cộ' and value == '1':
                check_xeco = True
            if level2 == 'thời trang' and value == '1':
                check_thoitrang = True
            if level2 == 'nhạc' and value == '1':
                check_nhac = True
            if level2 == ('thực phẩm' or 'thực phẩm giàu dinh dưỡng') and value == '1':
                check_thucpham = True
            if level2 == 'phim' and value == '1':
                check_phim = True
            if level2 == 'làm đẹp' and value == '1':
                check_lamdep = True



        if check_nganhang:
            global is_nganhang
            is_nganhang += 1
        if check_baohiem:
            global is_baohiem
            is_baohiem += 1
        if check_vay:
            global is_vay
            is_vay += 1
        if check_the:
            global is_the
            is_the += 1
        if check_dautu:
            global is_dautu
            is_dautu += 1
        if check_suckhoe:
            global is_suckhoe
            is_suckhoe += 1
        if check_giaoduc:
            global is_giaoduc
            is_giaoduc += 1
        if check_dulichnuocngoai:
            global is_dulichnuocngoai
            is_dulichnuocngoai += 1
        if check_dulichtrongnuoc:
            global is_dulichtrongnuoc
            is_dulichtrongnuoc += 1
        if check_duhoc:
            global is_duhoc
            is_duhoc += 1
        if check_thethao:
            global is_thethao
            is_thethao += 1
        if check_anuong:
            global is_anuong
            is_anuong += 1
        if check_khoahoc:
            global is_khoahoc
            is_khoahoc += 1
        if check_nghethuat:
            global is_nghethuat
            is_nghethuat += 1
        if check_game:
            global is_game
            is_game += 1
        if check_sothichsang:
            global is_sothichsang
            is_sothichsang += 1
        if check_xeco:
            global is_xeco
            is_xeco += 1
        if check_thoitrang:
            global is_thoitrang
            is_thoitrang += 1
        if check_nhac:
            global is_nhac
            is_nhac += 1
        if check_thucpham:
            global is_thucpham
            is_thucpham += 1
        if check_phim:
            global is_phim
            is_phim += 1
        if check_lamdep:
            global is_lamdep
            is_lamdep += 1
        wb.close()
        return dictObj, id

# if __name__ == "__main__":
#     df = pd.read_excel('cellphone 1000 truong.xlsx', converters={'cellphone': str})
#     list_phone = []
#     for index, row in df.iterrows():
#         if row['count field'] != 'error':
#             list_phone.append(row['cellphone'])
#     print(list_phone)
#     field = ['tuổi', 'giới tính', 'tình trạng hôn nhân', 'con cái', 'nghề nghiệp', 'trình độ học vấn', 'ngôn ngữ',
#              'quê quán', 'nơi ở hiện tại', 'sở hữu', 'ngân hàng', 'bảo hiểm', 'vay', 'thẻ', 'đầu tư', 'sức khỏe',
#              'giáo dục', 'du lịch nước ngoài', 'du lịch trong nước', 'du học', 'thể thao', 'ăn uống',
#              'khoa học công nghệ', 'nghệ thuật', 'game', 'sở thích sang', 'xe cộ', 'thời trang', 'nhạc', 'thực phẩm', 'phim', 'làm đẹp']
#     do_phu = []
#
#     def create_field():
#         wb = openpyxl.load_workbook('demographic_behaviorKeywords.xlsx')
#         user_field = []
#         Sheet1 = wb['Sheet1']
#         len1 = 0
#         len2 = 0
#         len3 = 0
#         len4 = 0
#         len5 = 0
#         len6 = 0
#         lv3 = []
#         level1 = ''
#         level2 = ''
#         level3 = ''
#         level4 = ''
#         level5 = ''
#         level6 = ''
#         for i in range(3, 1215):
#             cellNameA = f'A{i}'
#             cellNameB = f'B{i}'
#             cellNameC = f'C{i}'
#             cellNameD = f'D{i}'
#             cellNameE = f'E{i}'
#             cellNameF = f'F{i}'
#             cellNameG = f'G{i}'
#
#             cellDataA = Sheet1[cellNameA].value
#             cellDataB = Sheet1[cellNameB].value
#             cellDataC = Sheet1[cellNameC].value
#             cellDataD = Sheet1[cellNameD].value
#             cellDataE = Sheet1[cellNameE].value
#             cellDataF = Sheet1[cellNameF].value
#             cellDataG = Sheet1[cellNameG].value
#             if cellDataA == 'BEHAVIOR INDICATORS':
#                 continue
#             if cellDataB is None and cellDataC is None and cellDataD is None and cellDataE is None and cellDataF is None and cellDataG is None:
#                 continue
#             field = ''
#             if cellDataB is not None:
#
#                 level1 = cellDataB
#                 if cellDataC is None:
#                     field = level1
#                     len1 += 1
#
#             if cellDataC is not None:
#
#                 level2 = cellDataC
#                 if cellDataD is None:
#                     field =level1 + "." + level2
#                     len2 += 1
#
#             if cellDataD is not None:
#
#                 level3 = cellDataD
#                 if cellDataE is None:
#                     field = level1 + "." + level2 + "." + level3
#                     len3 += 1
#             if cellDataE is not None:
#
#                 level4 = cellDataE
#                 if cellDataF is None:
#                     field = level1 + "." + level2 + "." + level3 + "." + level4
#                     len4 += 1
#
#             if cellDataF is not None:
#
#                 level5 = cellDataF
#                 if cellDataG is None:
#                     field = level1 + "." + level2 + "." + level3 + "." + level4 + "." + level5
#                     len5 += 1
#                 else:
#                     field += "."
#             if cellDataG is not None:
#                 len6 += 1
#                 level6 = cellDataG
#                 field = level1 + "." + level2 + "." + level3 + "." + level4 + "." + level5 + "." + level6
#
#             user_field.append(field)
#         # print(len1, len2, len3, len4, len5, len6)
#         # print(lv3)
#         return user_field
#
#     df_get = DataFrame({'field': create_field()})
#
#     dict1 = 0
#     dict2 = 0
#     dict3 = 0
#     dict4 = 0
#     dict5 = 0
#     dict6 = 0
#
#     total_time = 0
#
#     for i in list_phone:
#         start_time = time.time()
#
#         try:
#             dictObj, id = API2().get_all(i)
#
#             user_result = []
#             for key1, value1 in dictObj.items():
#                 # print(key1)
#                 # dict1 = len(value1)
#                 # print(value1)
#
#                 if type(value1) == dict:
#                     # print(value1)
#                     for key2, value2 in value1.items():
#
#                         if type(value2) == dict:
#                             # print(value2 == '1')
#                             for key3, value3 in value2.items():
#                                 if type(value3) == dict:
#                                     for key4, value4 in value3.items():
#
#
#                                         if type(value4) == dict:
#                                             for key5, value5 in value4.items():
#
#
#                                                 if type(value5) == dict:
#                                                     for key6, value6 in value5.items():
#                                                         dict6 += 1
#                                                         user_result.append(value6)
#                                                 else:
#                                                     dict5 += 1
#                                                     user_result.append(value5)
#                                         else:
#                                             dict4 += 1
#                                             user_result.append(value4)
#                                 else:
#                                     dict3 += 1
#                                     user_result.append(value3)
#                         else:
#                             dict2 += 1
#                             user_result.append(value2)
#                 else:
#                     # print(value1 == '')
#                     dict1 += 1
#                     user_result.append(value1)
#             # print(k3)
#             len(user_in_post += 1
#             print(len(user_in_post)
#             # print(is_xeco/len(user_in_post)
#             # print(user_result)
#             # print(dictObj)
#         except:
#             continue
#         df_get[i + '( facebook.com/' + id + ' )'] = user_result
#         run_time_user = time.time() - start_time
#         total_time += run_time_user
#         print(run_time_user)
#     print("avg time / person: " + str(total_time/len(user_in_post))
#
#
#     do_phu.append(is_age / len(user_in_post * 100)
#     do_phu.append(is_gender / len(user_in_post * 100)
#     do_phu.append(is_relationship / len(user_in_post * 100)
#     do_phu.append(is_child / len(user_in_post * 100)
#     do_phu.append(is_job / len(user_in_post * 100)
#     do_phu.append(is_edu / len(user_in_post * 100)
#     do_phu.append(is_language / len(user_in_post * 100)
#     do_phu.append(is_hometown / len(user_in_post * 100)
#     do_phu.append(is_location / len(user_in_post * 100)
#     do_phu.append(is_own / len(user_in_post * 100)
#     do_phu.append(is_nganhang / len(user_in_post * 100)
#     do_phu.append(is_baohiem / len(user_in_post * 100)
#     do_phu.append(is_vay / len(user_in_post * 100)
#     do_phu.append(is_the / len(user_in_post * 100)
#     do_phu.append(is_dautu / len(user_in_post * 100)
#     do_phu.append(is_suckhoe / len(user_in_post * 100)
#     do_phu.append(is_giaoduc / len(user_in_post * 100)
#     do_phu.append(is_dulichnuocngoai / len(user_in_post * 100)
#     do_phu.append(is_dulichtrongnuoc / len(user_in_post * 100)
#     do_phu.append(is_duhoc / len(user_in_post * 100)
#     do_phu.append(is_thethao / len(user_in_post * 100)
#     do_phu.append(is_anuong / len(user_in_post * 100)
#     do_phu.append(is_khoahoc / len(user_in_post * 100)
#     do_phu.append(is_nghethuat / len(user_in_post * 100)
#     do_phu.append(is_game / len(user_in_post * 100)
#     do_phu.append(is_sothichsang / len(user_in_post * 100)
#     do_phu.append(is_xeco / len(user_in_post * 100)
#     do_phu.append(is_thoitrang / len(user_in_post * 100)
#     do_phu.append(is_nhac / len(user_in_post * 100)
#     do_phu.append(is_thucpham / len(user_in_post * 100)
#     do_phu.append(is_phim / len(user_in_post * 100)
#     do_phu.append(is_lamdep / len(user_in_post * 100)
#     df = DataFrame({'field': field, 'độ phủ(%)': do_phu})
#     df.to_excel(r'check do phu cac truong.xlsx', encoding='utf-8')
#     df_get.to_excel(r'get 1000 truong.xlsx', encoding='utf-8')




# def get_infor_and_post_byuser(userId):
#     res_user = es.search(index="user_cogroup_index", body={
#         "query": {
#             "match_phrase": {
#                 "id": userId
#             }
#         },
#     })
#     infor = res_user["hits"]["hits"][0]['_source']
#     post = []
#     res_post = es.search(index="post_index", body={
#         "size": 500000,
#         "query": {
#             "match_phrase": {
#                 "sourceId": userId
#             }
#         },
#         '_source': [
#             # "shortformDate", "message"
#         ],
#         # "sort": [
#         #     {"shortformDate": {"order": "asc"}}
#         # ]
#     })
#     for i in res_post["hits"]["hits"]:
#         post.append(i["_source"])
#     res_page = es2.search
#     user = {
#         "infor": infor,
#         "infor_post": post
#     }
#     return user
#
#
# res_post = es.search(index="post_index", body={
#         "size": 500000,
#         "query": {
#             "match_all": {
#             }
#         },
#         '_source': [
#             "sourceId"
#         ],
#         # "sort": [
#         #     {"shortformDate": {"order": "asc"}}
#         # ]
#     })
# user_in_post = set()
# for i in res_post["hits"]["hits"]:
#     user_in_post.add(i["_source"]["sourceId"])
#
# def create_field():
#     wb = openpyxl.load_workbook('demographic_behaviorKeywords.xlsx')
#     user_field = []
#     Sheet1 = wb['Sheet1']
#     len1 = 0
#     len2 = 0
#     len3 = 0
#     len4 = 0
#     len5 = 0
#     len6 = 0
#     lv3 = []
#     level1 = ''
#     level2 = ''
#     level3 = ''
#     level4 = ''
#     level5 = ''
#     level6 = ''
#     for i in range(3, 1215):
#         cellNameA = f'A{i}'
#         cellNameB = f'B{i}'
#         cellNameC = f'C{i}'
#         cellNameD = f'D{i}'
#         cellNameE = f'E{i}'
#         cellNameF = f'F{i}'
#         cellNameG = f'G{i}'
#
#         cellDataA = Sheet1[cellNameA].value
#         cellDataB = Sheet1[cellNameB].value
#         cellDataC = Sheet1[cellNameC].value
#         cellDataD = Sheet1[cellNameD].value
#         cellDataE = Sheet1[cellNameE].value
#         cellDataF = Sheet1[cellNameF].value
#         cellDataG = Sheet1[cellNameG].value
#         if cellDataA == 'BEHAVIOR INDICATORS':
#             continue
#         if cellDataB is None and cellDataC is None and cellDataD is None and cellDataE is None and cellDataF is None and cellDataG is None:
#             continue
#         field = ''
#         if cellDataB is not None:
#
#             level1 = cellDataB
#             if cellDataC is None:
#                 field = level1
#                 len1 += 1
#
#         if cellDataC is not None:
#
#             level2 = cellDataC
#             if cellDataD is None:
#                 field = level1 + "." + level2
#                 len2 += 1
#
#         if cellDataD is not None:
#
#             level3 = cellDataD
#             if cellDataE is None:
#                 field = level1 + "." + level2 + "." + level3
#                 len3 += 1
#         if cellDataE is not None:
#
#             level4 = cellDataE
#             if cellDataF is None:
#                 field = level1 + "." + level2 + "." + level3 + "." + level4
#                 len4 += 1
#
#         if cellDataF is not None:
#
#             level5 = cellDataF
#             if cellDataG is None:
#                 field = level1 + "." + level2 + "." + level3 + "." + level4 + "." + level5
#                 len5 += 1
#             else:
#                 field += "."
#         if cellDataG is not None:
#             len6 += 1
#             level6 = cellDataG
#             field = level1 + "." + level2 + "." + level3 + "." + level4 + "." + level5 + "." + level6
#
#         user_field.append(field)
#     # print(len1, len2, len3, len4, len5, len6)
#     # print(lv3)
#     return user_field
#
# field = ['tuổi', 'giới tính', 'tình trạng hôn nhân', 'con cái', 'nghề nghiệp', 'trình độ học vấn', 'ngôn ngữ',
#              'quê quán', 'nơi ở hiện tại', 'sở hữu', 'ngân hàng', 'bảo hiểm', 'vay', 'thẻ', 'đầu tư', 'sức khỏe',
#              'giáo dục', 'du lịch nước ngoài', 'du lịch trong nước', 'du học', 'thể thao', 'ăn uống',
#              'khoa học công nghệ', 'nghệ thuật', 'game', 'sở thích sang', 'xe cộ', 'thời trang', 'nhạc', 'thực phẩm', 'phim', 'làm đẹp']
# do_phu = []
# df_get = DataFrame({'field': create_field()})
# count = 0
#
# for user_id in list(user_in_post):
#     start_time = time.time()
#     user_data = get_infor_and_post_byuser(user_id)
#     try:
#         dictObj, id = API2().get_all(user_id, user_data)
#         count += 1
#         print(count)
#         user_result = []
#         for key1, value1 in dictObj.items():
#             # print(key1)
#             # dict1 = len(value1)
#             # print(value1)
#
#             if type(value1) == dict:
#                 # print(value1)
#                 for key2, value2 in value1.items():
#
#                     if type(value2) == dict:
#                         # print(value2 == '1')
#                         for key3, value3 in value2.items():
#                             if type(value3) == dict:
#                                 for key4, value4 in value3.items():
#
#                                     if type(value4) == dict:
#                                         for key5, value5 in value4.items():
#
#                                             if type(value5) == dict:
#                                                 for key6, value6 in value5.items():
#                                                     user_result.append(value6)
#                                             else:
#                                                 user_result.append(value5)
#                                     else:
#                                         user_result.append(value4)
#                             else:
#                                 user_result.append(value3)
#                     else:
#                         user_result.append(value2)
#             else:
#                 # print(value1 == '')
#                 user_result.append(value1)
#         # print(k3)
#         # print(is_xeco/len(user_in_post)
#         # print(user_result)
#         # print(dictObj)
#     except:
#         continue
#     df_get['( facebook.com/' + id + ' )'] = user_result
#     run_time_user = time.time() - start_time
#     print(run_time_user)
# do_phu.append(is_age / len(user_in_post) * 100)
# do_phu.append(is_gender / len(user_in_post) * 100)
# do_phu.append(is_relationship / len(user_in_post) * 100)
# do_phu.append(is_child / len(user_in_post) * 100)
# do_phu.append(is_job / len(user_in_post) * 100)
# do_phu.append(is_edu / len(user_in_post) * 100)
# do_phu.append(is_language / len(user_in_post) * 100)
# do_phu.append(is_hometown / len(user_in_post) * 100)
# do_phu.append(is_location / len(user_in_post) * 100)
# do_phu.append(is_own / len(user_in_post) * 100)
# do_phu.append(is_nganhang / len(user_in_post) * 100)
# do_phu.append(is_baohiem / len(user_in_post) * 100)
# do_phu.append(is_vay / len(user_in_post) * 100)
# do_phu.append(is_the / len(user_in_post) * 100)
# do_phu.append(is_dautu / len(user_in_post) * 100)
# do_phu.append(is_suckhoe / len(user_in_post) * 100)
# do_phu.append(is_giaoduc / len(user_in_post) * 100)
# do_phu.append(is_dulichnuocngoai / len(user_in_post) * 100)
# do_phu.append(is_dulichtrongnuoc / len(user_in_post) * 100)
# do_phu.append(is_duhoc / len(user_in_post) * 100)
# do_phu.append(is_thethao / len(user_in_post) * 100)
# do_phu.append(is_anuong / len(user_in_post) * 100)
# do_phu.append(is_khoahoc / len(user_in_post) * 100)
# do_phu.append(is_nghethuat / len(user_in_post) * 100)
# do_phu.append(is_game / len(user_in_post) * 100)
# do_phu.append(is_sothichsang / len(user_in_post) * 100)
# do_phu.append(is_xeco / len(user_in_post) * 100)
# do_phu.append(is_thoitrang / len(user_in_post) * 100)
# do_phu.append(is_nhac / len(user_in_post) * 100)
# do_phu.append(is_thucpham / len(user_in_post) * 100)
# do_phu.append(is_phim / len(user_in_post) * 100)
# do_phu.append(is_lamdep / len(user_in_post) * 100)
# df = DataFrame({'field': field, 'độ phủ(%)': do_phu})
# df.to_excel(r'check do phu cac truong.xlsx', encoding='utf-8')
# df_get.to_excel(r'get 1000 truong.xlsx', encoding='utf-8')



